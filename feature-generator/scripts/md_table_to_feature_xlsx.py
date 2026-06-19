#!/usr/bin/env python3
"""Create a styled feature-spec XLSX from a Markdown feature table."""

from __future__ import annotations

import argparse
import html
import math
import re
import shutil
import unicodedata
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent
TEMPLATE_PATH = SKILL_DIR / "assets" / "feature_xlsx_template.xlsx"

HEADERS = [
    "1depth",
    "2depth",
    "3depth",
    "요구사항 ID",
    "요구사항명",
    "요청목적",
    "기능 요구사항",
    "프로세스 요구사항",
    "화면 요구사항",
    "보안 요구사항",
    "데이터 요구사항",
]

COLUMN_WIDTHS = {
    "A": 20.0,
    "B": 24.0,
    "C": 30.0,
    "D": 21.6640625,
    "E": 42.83203125,
    "F": 39.1640625,
    "G": 74.1640625,
    "H": 71.5,
    "I": 63.83203125,
    "J": 48.1640625,
    "K": 55.5,
}

HEADER_ROW = 5
DATA_START_ROW = 6
MAX_COL = len(HEADERS)
ODD_FILL = PatternFill(fill_type="solid", fgColor="FFFFFFFF")
EVEN_FILL = PatternFill(fill_type="solid", fgColor="FFF2F2F2")
MIN_BODY_HEIGHT = 93.75
HEADER_HEIGHT = 32.0


def split_md_row(line: str) -> list[str]:
    raw = line.strip().strip("|")
    cells = re.split(r"(?<!\\)\|", raw)
    return [clean_cell(cell) for cell in cells]


def clean_cell(value: str) -> str:
    value = value.strip().replace(r"\|", "|")
    value = value.replace("<br>", "\n").replace("&lt;br&gt;", "\n")
    return html.unescape(value)


def is_separator_row(row: list[str]) -> bool:
    return all(re.fullmatch(r"\s*:?-{3,}:?\s*", cell or "") for cell in row)


def extract_markdown_table(md_text: str) -> list[list[str]]:
    lines = md_text.splitlines()
    table_blocks: list[list[str]] = []
    current: list[str] = []
    for line in lines:
        if line.startswith("|"):
            current.append(line)
        elif current:
            table_blocks.append(current)
            current = []
    if current:
        table_blocks.append(current)

    for block in table_blocks:
        rows = [split_md_row(line) for line in block]
        rows = [row for row in rows if row and not is_separator_row(row)]
        if rows and rows[0][:3] == ["1depth", "2depth", "3depth"]:
            return normalize_rows(rows)

    raise ValueError("No feature table with 1depth/2depth/3depth headers found")


def normalize_rows(rows: list[list[str]]) -> list[list[str]]:
    header = rows[0]
    index_by_header = {name: i for i, name in enumerate(header)}
    normalized = [HEADERS]
    for row in rows[1:]:
        normalized.append([
            row[index_by_header[name]] if name in index_by_header and index_by_header[name] < len(row) else ""
            for name in HEADERS
        ])
    return normalized


def display_width(text: str) -> int:
    width = 0
    for char in text:
        if unicodedata.east_asian_width(char) in {"F", "W"}:
            width += 2
        else:
            width += 1
    return width


def estimate_wrapped_lines(text: str, column_width: float) -> int:
    if not text:
        return 1
    usable_width = max(int(column_width * 1.35), 8)
    total = 0
    for part in str(text).splitlines() or [""]:
        total += max(1, math.ceil(display_width(part) / usable_width))
    return total


def estimate_row_height(row: list[str]) -> float:
    max_lines = 1
    for idx, value in enumerate(row, start=1):
        column_letter = chr(ord("A") + idx - 1)
        max_lines = max(max_lines, estimate_wrapped_lines(value, COLUMN_WIDTHS[column_letter]))
    estimated = max_lines * 18 + 22
    return max(MIN_BODY_HEIGHT, float(estimated))


def copy_cell_style(src, dst) -> None:
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = "@"
    dst.protection = copy(src.protection)
    dst.quotePrefix = True


def force_text_cell(cell, value: str) -> None:
    cell.value = "" if value is None else str(value)
    cell.data_type = "s"
    cell.number_format = "@"
    cell.quotePrefix = True


def build_workbook(rows: list[list[str]], output_path: Path) -> None:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Missing template asset: {TEMPLATE_PATH}")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(TEMPLATE_PATH, output_path)

    wb = load_workbook(output_path)
    ws = wb.active
    ws.title = "기능명세서"
    ws.data_validations.dataValidation = []

    existing_data_rows = max(ws.max_row - DATA_START_ROW + 1, 0)
    required_data_rows = max(len(rows) - 1, 0)
    if existing_data_rows < required_data_rows:
        ws.insert_rows(DATA_START_ROW + existing_data_rows, required_data_rows - existing_data_rows)
    elif existing_data_rows > required_data_rows:
        ws.delete_rows(DATA_START_ROW + required_data_rows, existing_data_rows - required_data_rows)

    # Header
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(HEADER_ROW, col_idx)
        copy_cell_style(ws.cell(HEADER_ROW, col_idx), cell)
        force_text_cell(cell, header)
    ws.row_dimensions[HEADER_ROW].height = HEADER_HEIGHT

    # Body
    for r_offset, row in enumerate(rows[1:], start=0):
        row_idx = DATA_START_ROW + r_offset
        fill = EVEN_FILL if (r_offset + 1) % 2 == 0 else ODD_FILL
        for col_idx in range(1, MAX_COL + 1):
            cell = ws.cell(row_idx, col_idx)
            copy_cell_style(ws.cell(DATA_START_ROW, col_idx), cell)
            cell.fill = copy(fill)
            force_text_cell(cell, row[col_idx - 1] if col_idx - 1 < len(row) else "")
        ws.row_dimensions[row_idx].height = estimate_row_height(row)

    # Clear cells outside A:K in the table area if any survived from template history.
    if ws.max_column > MAX_COL:
        ws.delete_cols(MAX_COL + 1, ws.max_column - MAX_COL)

    for column, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[column].width = width

    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    wb.save(output_path)


def scan_for_errors(path: Path) -> None:
    wb = load_workbook(path, data_only=False)
    for ws in wb.worksheets:
        if ws.data_validations.dataValidation:
            raise ValueError("Data validation restrictions remain")
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    if cell.value.startswith("="):
                        raise ValueError(f"Formula-like value found at {ws.title}!{cell.coordinate}")
                    if cell.value in {"#NAME?", "#VALUE!", "#REF!", "#DIV/0!", "#N/A"}:
                        raise ValueError(f"Excel error value found at {ws.title}!{cell.coordinate}")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("input_md", type=Path)
    parser.add_argument("output_xlsx", type=Path)
    args = parser.parse_args()

    rows = extract_markdown_table(args.input_md.read_text(encoding="utf-8"))
    build_workbook(rows, args.output_xlsx)
    scan_for_errors(args.output_xlsx)
    print(args.output_xlsx)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
